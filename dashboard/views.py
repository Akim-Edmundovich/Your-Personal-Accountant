from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib.auth.decorators import login_required
from django.template.loader import render_to_string
from django.db.models import Sum
from django.utils import timezone

import io
import xlsxwriter
from openpyxl import Workbook
from datetime import timedelta, date

from dashboard.resources import TransactionResource
from transactions.models import *


@login_required
def test(request):
    return render(request, 'index.html')


@login_required
def dashboard(request):
    return render(request, 'dashboard/dashboard.html')


@login_required
def list_transactions(request, category: str):
    transactions = Transaction.objects.filter(user=request.user,
                                              category__name=category).select_related(
        'category').select_related('subcategory')

    order_by_date = transactions.order_by('-created_at')
    order_by_amount = transactions.order_by('-amount')

    context = {'transactions': transactions,
               'order_by_date': order_by_date,
               'order_by_amount': order_by_amount, }
    return render(request, 'dashboard/list_transactions.html', context)


@login_required
def detail_transaction(request, pk: int):
    transaction = Transaction.objects.get(id=pk)
    categories = Category.objects.all()
    subcategories = Subcategory.objects.all()

    context = {
        'transaction': transaction,
        'categories': categories,
        'subcategories': subcategories,
    }
    return render(request, 'transaction/detail_transaction.html', context)


@login_required
def update_transaction(request, pk: int):
    transaction = Transaction.objects.get(id=pk, user=request.user)

    if request.method == 'POST':
        type = request.POST.get('transaction_type')
        category_id = request.POST.get('category')
        subcategory_id = request.POST.get('subcategory')
        amount = request.POST.get('amount')
        quantity = request.POST.get('quantity')
        quantity_type = request.POST.get('quantity_type')
        description = request.POST.get('description')
        created_at = request.POST.get('created_at')

        print(f'{type} - {category_id} - {subcategory_id} - {amount} - '
              f'{quantity} - {quantity_type} - {description} - {created_at}')

        try:
            Transaction.objects.filter(id=pk).update(
                type=type,
                amount=amount,
                category=category_id,
                subcategory=subcategory_id,
                quantity=quantity,
                quantity_type=quantity_type,
                description=description,
                created_at=created_at
            )

            print(
                f'Category - {transaction.category}, '
                f'subcategory - {transaction.subcategory}, '
                f'amount - {transaction.amount}')

            return redirect('dashboard:list_transactions')

        except Exception as e:
            print(f'Error while updating transaction: {e}')

    return redirect('dashboard:list_transactions')


@login_required
def delete_transaction(request, pk: int):
    transaction = get_object_or_404(Transaction, id=pk)

    if request.method == 'POST':
        try:
            transaction.delete()
            return redirect('dashboard:list_transactions', transaction.category)

        except transaction.DoesNotExist as e:
            print(f'Cannot delete transaction. Detail: {e}')

    context = {'transaction': transaction}
    return render(request, 'transaction/delete_transaction.html', context)


# ------------------------------------------------------

def convert_decimal_to_str(transactions):
    return (
        {category: str(amount) for category, amount in transactions.items()})


def calculate_sum_by_category(user,
                              transaction_type: str,
                              filter_type: str,
                              start_date: date = None,
                              end_date: date = None):
    categories = Category.objects.all()
    category_and_sum = {}
    today = timezone.now().date()

    for category in categories:
        date_filter = {}

        if filter_type == f'day_{transaction_type}':
            date_filter = {'created_at': start_date}

        elif (filter_type == f'week_{transaction_type}' or
              filter_type == f'period_{transaction_type}'):
            if start_date and end_date:
                date_filter = {'created_at__range': [start_date, end_date]}

        elif filter_type == f'month_{transaction_type}':
            date_filter = {'created_at__year': today.year,
                           'created_at__month': today.month}

        elif filter_type == f'year_{transaction_type}' and start_date:
            date_filter = {'created_at__year': today.year}

        summ = Transaction.objects.filter(type=transaction_type,
                                          category=category,
                                          user=user,
                                          **date_filter,
                                          ).aggregate(Sum('amount'))
        total_amount = summ['amount__sum']
        if total_amount is not None:
            category_and_sum[category.name] = total_amount

    return category_and_sum


@login_required
def expenses_filter_transactions(request, filter_type: str):
    today = timezone.now().date()
    expenses = Transaction.objects.none()

    if filter_type == 'day_expense':
        expenses = calculate_sum_by_category(request.user,
                                             'expense',
                                             'day_expense',
                                             start_date=today)

    elif filter_type == 'week_expense':
        start_date = today - timedelta(days=today.weekday())
        end_date = start_date + timedelta(days=6)
        expenses = calculate_sum_by_category(request.user,
                                             'expense',
                                             'week_expense',
                                             start_date=start_date,
                                             end_date=end_date)
    elif filter_type == 'month_expense':
        expenses = calculate_sum_by_category(request.user,
                                             'expense',
                                             'month_expense',
                                             )

    elif filter_type == 'year_expense':
        expenses = calculate_sum_by_category(request.user,
                                             'expense',
                                             'year_expense',
                                             start_date=today)
    elif filter_type == 'period_expense':
        date_range = request.GET.get('date_range')
        if date_range:
            start_date, end_date = date_range.split(',')
            expenses = calculate_sum_by_category(request.user,
                                                 'expense',
                                                 'period_expense',
                                                 start_date=start_date,
                                                 end_date=end_date)

    return convert_decimal_to_str(expenses)


@login_required
def to_html_format_expenses_filter(request, filter_type):
    expenses_data = expenses_filter_transactions(request, filter_type)
    html = render_to_string(
        'dashboard/filter_transactions/expenses_filter_transaction.html',
        {'expenses': expenses_data})

    return JsonResponse({'html': html})


@login_required
def incomes_filter_transactions(request, filter_type: str):
    today = timezone.now().date()
    incomes = Transaction.objects.none()

    if filter_type == 'day_income':
        incomes = calculate_sum_by_category(request.user,
                                            'income',
                                            'day_income',
                                            start_date=today)

    elif filter_type == 'week_income':
        start_date = today - timedelta(days=today.weekday())
        end_date = start_date + timedelta(days=6)
        incomes = calculate_sum_by_category(request.user,
                                            'income',
                                            'week_income',
                                            start_date=start_date,
                                            end_date=end_date)
    elif filter_type == 'month_income':
        incomes = calculate_sum_by_category(request.user,
                                            'income',
                                            'month_income',
                                            )

    elif filter_type == 'year_income':
        incomes = calculate_sum_by_category(request.user,
                                            'income',
                                            'year_income',
                                            start_date=today)
    elif filter_type == 'period_income':
        date_range = request.GET.get('date_range')
        if date_range:
            start_date, end_date = date_range.split(',')
            incomes = calculate_sum_by_category(request.user,
                                                'income',
                                                'period_income',
                                                start_date=start_date,
                                                end_date=end_date)

    return convert_decimal_to_str(incomes)


@login_required
def to_html_format_incomes_filter(request, filter_type):
    incomes_data = incomes_filter_transactions(request, filter_type)
    html = render_to_string(
        'dashboard/filter_transactions/incomes_filter_transaction.html',
        {'incomes': incomes_data}
    )

    return JsonResponse({'html': html})


@login_required
def export_order_by_category(request, file_format: str):
    transactions = Transaction.objects.all().order_by('category__name')

    resource = TransactionResource()
    dataset = resource.export(transactions)

    if file_format == 'csv':
        response = HttpResponse(
            dataset.export('csv', delimiter=';').encode('utf-8-sig'),
            content_type=f'text/{file_format}; charset=utf-8')
        response[
            'Content-Disposition'] = f'attachment; filename="transactions.{file_format}"'
        return response
    elif file_format == 'xlsx':
        response = HttpResponse(
            dataset.export('xlsx'),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response[
            'Content-Disposition'] = 'attachment; filename="transactions.xlsx"'
        return response


@login_required
def to_xlsx_openpyxl(request, filter_type):
    # Создать файл в памяти
    output = io.BytesIO()
    wb = Workbook()

    wb.remove(wb.active)

    ws1 = wb.create_sheet('Expense')
    ws2 = wb.create_sheet('Income')

    expense = expenses_filter_transactions(request, filter_type)
    income = incomes_filter_transactions(request, filter_type)

    # Записать Expense
    for category, amount in expense.items():
        ws1.append([category, amount])

    # Записать Income
    for category, amount in income.items():
        ws2.append([category, amount])

    # Сохранить книгу в объект памяти
    wb.save(output)
    wb.close()
    output.seek(0)

    response = HttpResponse(output,
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response[
        'Content-Disposition'] = 'attachment; filename="expenses_report.xlsx"'

    return response


@login_required
def to_xlsx_formatter(request, filter_type):
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet()

    expenses = expenses_filter_transactions(request, filter_type)

    bold = workbook.add_format({'bold': 1})

    worksheet.set_column(1, 1, 15)

    worksheet.write('A1', 'Category', bold)
    worksheet.write('B1', 'Amount', bold)

    row = 0
    for category, amount in expenses.items():
        worksheet.write(row, 0, category)
        worksheet.write(row, 1, amount)
        row += 1

    workbook.close()

    output.seek(0)
    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response[
        'Content-Disposition'] = 'attachment; filename="expenses_report.xlsx"'

    return response
